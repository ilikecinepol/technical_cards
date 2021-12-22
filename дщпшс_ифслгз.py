from docxtpl import DocxTemplate
import time
import openpyxl
import datetime
import generator

def start_generator():


    pattern = str(generator.reference_file)
    ktp = str(generator.ktp_file)
    group = str(generator.classmate)
    lesson = str(generator.lesson)
    hours = 0

    n = 64
    day = time.localtime()
    d = day[2]
    # print(type(d))

    themes_list = []
    date_dict = []
    lessons = {}
    number_list = []

    wb = openpyxl.load_workbook(ktp, data_only=True)
    sheet = wb.active

    name = str(generator.teacher)

    hours = sheet['F5'].value



    # print(hours)
    def input_hours(hours):
        ''' Данная функция получает начальное и конечное значение уроков из ячейки F5'''
        hours = hours.split()
        # print(hours)
        if len(hours) == 1:
            a = hours[0]
            for x in a:
                if x not in '123456789':
                    sep = a.index(x)

            h1 = [x for x in a if a.index(x) < sep]
            h1 = int(''.join(h1))
            h2 = [x for x in a if a.index(x) > sep]
            h2 = int(''.join(h2))
        else:
            h1 = int(hours[0])
            h2 = int(hours[2])

        if h1 > h2:
            h1 = h1 + 1
            return h2, h1
        else:
            h2 = h2 + 1
            return h1, h2

    hours = input_hours(hours)
    # print(type(hours[0]), hours[1])
    number_of_lessons = int(hours[1]) - int(hours[0])

    def table(group, lesson, ):
        pattern = f'{lesson}{group}'


        wb = openpyxl.load_workbook(ktp, data_only=True)
        # печатаем список листов
        sheets = wb.sheetnames

        sheet = wb.active
        # print(sheet)
        # print(sheet['A1'])
        # print(sheet['A1'].value)
        # Перебираем темы
        for row in sheet[f'B{hours[0]}':f'B{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                # print(cellObj.value)
                themes_list.append(cellObj.value)
        # Перебираем номера уроков
        for row in sheet[f'A{hours[0]}':f'A{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                # print(cellObj.value)
                number_list.append(cellObj.value)
        # Перебираем даты
        for row in sheet[f'C{hours[0]}':f'C{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                print(type(cellObj.value))
                try:

                    if cellObj.value.month == 1:
                        result_mounth = 'января'
                    elif cellObj.value.month == 2:
                        result_mounth = 'февраля'
                    elif cellObj.value.month == 3:
                        result_mounth = 'марта'
                    elif cellObj.value.month == 4:
                        result_mounth = 'апреля'
                    elif cellObj.value.month == 5:
                        result_mounth = 'мая'
                    elif cellObj.value.month == 6:
                        result_mounth = 'июня'
                    elif cellObj.value.month == 7:
                        result_mounth = 'июля'
                    elif cellObj.value.month == 8:
                        result_mounth = 'августа'
                    elif cellObj.value.month == 9:
                        result_mounth = 'сентября'
                    elif cellObj.value.month == 10:
                        result_mounth = 'октября'
                    elif cellObj.value.month == 11:
                        result_mounth = 'ноября'
                    elif cellObj.value.month == 12:
                        result_mounth = 'декабря'
                    date_dict.append(f"{cellObj.value.day} {result_mounth} {cellObj.value.year}")
                except AttributeError:
                    pass
        lessons = {x: y for x in themes_list for y in date_dict}
        # print(date_dict)

        # cell_value = sheet.cell(column,row).value  # cell_value = sheet.cell(строка, столбец)
        # print(cell_value)
        # print(lessons)

    def automatic(pattern, group, lesson, hours):
        for x in range(number_of_lessons):
            doc = DocxTemplate(pattern)
            day = time.localtime()
            d = day[2]
            # n = 64
            context = {'name': name, 'lesson': lesson,
                       'theme': themes_list[x], 'date': date_dict[x], 'number': number_list[x], 'class': group}

            doc.render(context)
            doc.save(f"Техкарты готовые/Занятие_{number_list[x]}.docx")

            # n += 1
            if x % 2 == 0:
                d += 5
            else:
                d += 2

    table(pattern, group)
    automatic(pattern, group, lesson, hours)
    print(f'Добро пожаловать в программу "Генератор технологических карт!"')
    print("Генерируем для Вас файлы")
    for x in range(10):
        print("#####", end='')
        time.sleep(0.5)
    print()
    print("Вот и всё! Готовые файлы были сохранены в папку с программой!")

if __name__ == '__main__':
    start_generator()