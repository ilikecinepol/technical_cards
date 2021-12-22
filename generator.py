# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'card_generator_GUI.ui'
#
# Created by: PyQt5 UI code generator 5.15.4
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog
from PyQt5.uic import loadUi
import os
#import logic
import json

profiles = ''
teacher = ''
classmate = ''
lesson = ''
ktp_file = ''
reference_file = ''

'''teacher = 'Молотков М.А.'
classmate = '7'
lesson = 'Робототехника'
ktp_file = 'КТП.xlsx'
reference_file = 'Шаблон.docx' '''


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(799, 639)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 0, 801, 61))
        font = QtGui.QFont()
        font.setFamily("Chalk cyrillic freehand")
        font.setPointSize(38)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label.setFont(font)
        self.label.setStyleSheet("color: rgb(255, 255, 255);\n"
                                 "\n"
                                 "font: 38pt \"Chalk cyrillic freehand\";")
        self.label.setTextFormat(QtCore.Qt.PlainText)
        self.label.setObjectName("label")
        self.comboBox_profiles = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox_profiles.setGeometry(QtCore.QRect(20, 450, 251, 41))
        self.comboBox_profiles.setObjectName("comboBox_profiles")
        self.files = []
        for file in os.walk('profiles'):
            self.files.append(file)
        print(self.files)
        try:
            self.files = self.files[0][2]
        except:
            pass
        for i in range(len(self.files)):
            self.comboBox_profiles.addItem('')

        self.pushButton_confirm_profiles = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_confirm_profiles.setGeometry(QtCore.QRect(290, 450, 111, 41))
        self.pushButton_confirm_profiles.setObjectName("pushButton_confirm_profiles")

        self.pushButton_generate = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_generate.setGeometry(QtCore.QRect(430, 450, 330, 41))
        self.pushButton_generate.setObjectName("pushButton_generate")
        self.pushButton_generate.setStyleSheet("font: 17pt \"Chalk cyrillic freehand\";\n"
                                               "background-color: rgb(255, 0, 0);color: rgb(255, 255, 255);")

        self.pushButton_info = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_info.setGeometry(QtCore.QRect(650, 20, 111, 41))
        self.pushButton_info.setObjectName("pushButton_info")

        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 250, 371, 31))
        self.label_2.setStyleSheet("font: 13pt \"Chalk cyrillic freehand\";\n"
                                   "color: rgb(255, 255, 255);")
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(20, 330, 371, 31))
        self.label_3.setStyleSheet("font: 13pt \"Chalk cyrillic freehand\";\n"
                                   "color: rgb(255, 255, 255);")
        self.label_3.setObjectName("label_3")
        self.ktp_file_name = QtWidgets.QLineEdit(self.centralwidget)
        self.ktp_file_name.setGeometry(QtCore.QRect(20, 280, 251, 41))
        self.ktp_file_name.setStyleSheet("font: 10pt \"Arial\";")
        self.ktp_file_name.setObjectName("ktp_file_name")
        self.browse_ktp = QtWidgets.QPushButton(self.centralwidget)
        self.browse_ktp.setGeometry(QtCore.QRect(290, 280, 111, 41))
        self.browse_ktp.setObjectName("browse_ktp")
        self.reference_name = QtWidgets.QLineEdit(self.centralwidget)
        self.reference_name.setGeometry(QtCore.QRect(20, 370, 251, 41))
        self.reference_name.setStyleSheet("font: 10pt \"Arial\";")
        self.reference_name.setObjectName("reference_name")
        self.browse_reference = QtWidgets.QPushButton(self.centralwidget)
        self.browse_reference.setGeometry(QtCore.QRect(290, 370, 111, 41))
        self.browse_reference.setObjectName("browse_reference")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(0, 0, 800, 600))
        self.widget.setStyleSheet(
            "image: url(:/background/education-happy-teachers-day-jpg-pictures-to---picture-downloads-backgrounds.jpg);")
        self.widget.setObjectName("widget")
        self.textEdit = QtWidgets.QTextEdit(self.widget)
        self.textEdit.setGeometry(QtCore.QRect(430, 80, 331, 331))
        self.textEdit.setStyleSheet(
            "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
            "font: 20pt \"Chalk cyrillic freehand\";\n"
            "color: rgb(255, 255, 255);" "text-align: center")
        self.textEdit.setObjectName("textEdit")
        self.input_lesson = QtWidgets.QLineEdit(self.widget)
        self.input_lesson.setGeometry(QtCore.QRect(20, 70, 381, 51))
        self.input_lesson.setStyleSheet("font: 12pt \"Arial\";")
        self.input_lesson.setObjectName("input_lesson")
        self.input_teacher = QtWidgets.QLineEdit(self.widget)
        self.input_teacher.setGeometry(QtCore.QRect(20, 130, 381, 51))
        self.input_teacher.setStyleSheet("font: 12pt \"Arial\";")
        self.input_teacher.setObjectName("input_teacher")
        self.input_class = QtWidgets.QLineEdit(self.widget)
        self.input_class.setGeometry(QtCore.QRect(20, 190, 381, 51))
        self.input_class.setStyleSheet("font: 12pt \"Arial\";")
        self.input_class.setObjectName("input_class")
        self.widget.raise_()
        self.label.raise_()
        self.comboBox_profiles.raise_()
        self.pushButton_confirm_profiles.raise_()
        self.pushButton_generate.raise_()
        self.pushButton_info.raise_()

        self.label_2.raise_()
        self.label_3.raise_()
        self.ktp_file_name.raise_()
        self.browse_ktp.raise_()
        self.reference_name.raise_()
        self.browse_reference.raise_()
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 799, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.add_funktion()

    def add_funktion(self):
        self.browse_ktp.clicked.connect(lambda: self.browsfiles(self.ktp_file_name))
        self.browse_reference.clicked.connect(lambda: self.browsfiles(self.reference_name))
        # self.pushButton_confirm_profiles.clicked.connect(lambda: self.accept_profile(self.get_parametrs()))

        self.accept_profile(self.get_parametrs())
        self.pushButton_generate.clicked.connect(lambda: self.start_generator())
        self.pushButton_confirm_profiles.clicked.connect(lambda: self.accept_profile(self.get_parametrs()))
        self.pushButton_info.clicked.connect(lambda: self.open_info_file())
        # lambda: main.start_generator()

    def open_info_file(self):
        os.system(r"explorer.exe READ_ME.pdf")

    def printing_data(self):
        global teacher, lesson, classmate, ktp_file, reference_file
        return teacher, lesson, classmate, ktp_file, reference_file
        print(teacher, lesson, classmate, ktp_file, reference_file)

    def get_parametrs(self):
        global lesson, teacher, classmate
        profile_name = self.comboBox_profiles.currentText()
        lesson = self.input_lesson.text()
        teacher = self.input_teacher.text()
        classmate = self.input_class.text()
        self.out = f'''Название предмета: {lesson} \nФИО преподавателя: {teacher} \nКласс: {classmate}'''
        self.textEdit.setText(self.out)

        return profile_name, lesson, teacher, classmate

    def accept_profile(self, profile):
        global ktp_file, reference_file, profiles
        profiles = profile[0]
        self.get_parametrs()
        # print(profile[0])
        data = self.textEdit.toPlainText()
        ktp_file = self.ktp_file_name.text()
        reference_file = self.reference_name.text()
        # print(data)
        if profile[0] == 'Создать новый профиль.json':
            print('Создаём новый профиль')
            data = {
                "profile": {
                    "lesson": profile[1],
                    "teacher": profile[2],
                    "group": profile[3],
                    "ktp": ktp_file,
                    "reference": reference_file,
                }
            }
            with open(f"profiles/Создать новый профиль.json", "w", encoding='utf-8') as write_file:
                json.dump(data, write_file)
            with open(f"profiles/{profile[1]} - {profile[2]}.json", "w", encoding='utf-8') as write_file:
                json.dump(data, write_file)


            self.pushButton_confirm_profiles.clicked.connect(lambda: self.accept_profile(self.get_parametrs()))
        else:
            if profile[0] in self.files:
                with open(f'profiles/{profile[0]}', 'r') as file:
                    self.reading_data = json.load(file)

                self.out = f'''Название предмета: {self.reading_data['profile']['lesson']} \nФИО преподавателя: {self.reading_data['profile']['teacher']} \nКласс: {self.reading_data['profile']['group']}'''


                self.pushButton_confirm_profiles.clicked.connect(lambda: self.textEdit.setText(self.out))

    def browsfiles(self, object):
        fname = QFileDialog.getOpenFileName(None, 'Открыть файл', 'files')
        data = []
        for i in range(len(fname[0]) - 1, 0, -1):
            if fname[0][i] != '/':
                print(fname[0][i])
                data.append(fname[0][i])
            else:
                break
        data = data[::-1]
        data = ''.join(data)
        print(data)
        object.setText(data)

    def start_generator(self):
        global profiles
        self.printing_data()
        profiles = self.comboBox_profiles.currentText()
        start_generator_logic(profiles)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Генератор техкарт v1.0"))
        self.label.setText(_translate("MainWindow", "Генератор технологических карт"))

        self.comboBox_profiles.setItemText(0, _translate("MainWindow", "Создать новый профиль"))
        count = 0
        print(self.files)
        for i in self.files:
            self.comboBox_profiles.setItemText(count, i)
            count += 1

        self.pushButton_confirm_profiles.setText(_translate("MainWindow", "Применить"))
        self.pushButton_info.setText(_translate("MainWindow", "Инфо"))
        self.pushButton_generate.setText(_translate("MainWindow", "СГЕНЕРИРОВАТЬ КАРТЫ!"))
        self.label_2.setText(_translate("MainWindow", "Укажите файл с календарно-тематическим планированием"))
        self.label_3.setText(_translate("MainWindow", "Укажите файл с шаблоном технологической карты"))
        self.browse_ktp.setText(_translate("MainWindow", "Выбрать файл"))
        self.browse_reference.setText(_translate("MainWindow", "Выбрать файл"))
        self.input_lesson.setText(_translate("MainWindow", "Введите название предмета"))
        self.input_teacher.setText(_translate("MainWindow", "Введите ФИО преподавателя"))
        self.input_class.setText(_translate("MainWindow", "Введите класс"))


import resource_rc
def get_data():
    global lesson, teacher, classmate, ktp_file, reference_file
    return lesson, teacher, classmate, ktp_file, reference_file

from docxtpl import DocxTemplate
import time
import openpyxl
import datetime
#from generator import *

pattern = ''
ktp = ''
group = ''
lesson = ''
name = ''
hours = 0

def start_generator_logic(data):
    print('Название профиля:',data)

    with open(f'profiles/{data}', 'r') as file:
        reading_data = json.load(file)
    lesson = reading_data['profile']['lesson']
    name = reading_data['profile']['teacher']
    group = reading_data['profile']['group']
    ktp = 'files/' + reading_data['profile']['ktp']
    pattern = 'files/' + reading_data['profile']['reference']

    print(lesson, name, group, ktp, pattern)


    n = 64
    day = time.localtime()
    d = day[2]
    # print(type(d))

    themes_list = []
    date_dict = []
    lessons = {}
    number_list = []

    wb = openpyxl.load_workbook(ktp, data_only=True)
    sheet = wb.active

    hours = sheet['F5'].value

    # print(hours)
    def input_hours(hours):
        ''' Данная функция получает начальное и конечное значение уроков из ячейки F5'''
        hours = hours.split()
        # print(hours)
        if len(hours) == 1:
            a = hours[0]
            for x in a:
                if x not in '123456789':
                    sep = a.index(x)

            h1 = [x for x in a if a.index(x) < sep]
            h1 = int(''.join(h1))
            h2 = [x for x in a if a.index(x) > sep]
            h2 = int(''.join(h2))
        else:
            h1 = int(hours[0])
            h2 = int(hours[2])

        if h1 > h2:
            h1 = h1 + 1
            return h2, h1
        else:
            h2 = h2 + 1
            return h1, h2

    hours = input_hours(hours)
    # print(type(hours[0]), hours[1])
    number_of_lessons = int(hours[1]) - int(hours[0])

    def table(group, lesson, ):
        pattern = f'{lesson}{group}'

        wb = openpyxl.load_workbook(ktp, data_only=True)
        # печатаем список листов
        sheets = wb.sheetnames

        sheet = wb.active
        # print(sheet)
        # print(sheet['A1'])
        # print(sheet['A1'].value)
        # Перебираем темы
        for row in sheet[f'B{hours[0]}':f'B{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                # print(cellObj.value)
                themes_list.append(cellObj.value)
        # Перебираем номера уроков
        for row in sheet[f'A{hours[0]}':f'A{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                # print(cellObj.value)
                number_list.append(cellObj.value)
        # Перебираем даты
        for row in sheet[f'C{hours[0]}':f'C{hours[1]}']:
            for cellObj in row:
                if cellObj.value == None or cellObj.value == " ":
                    continue
                print(type(cellObj.value))
                try:

                    if cellObj.value.month == 1:
                        result_mounth = 'января'
                    elif cellObj.value.month == 2:
                        result_mounth = 'февраля'
                    elif cellObj.value.month == 3:
                        result_mounth = 'марта'
                    elif cellObj.value.month == 4:
                        result_mounth = 'апреля'
                    elif cellObj.value.month == 5:
                        result_mounth = 'мая'
                    elif cellObj.value.month == 6:
                        result_mounth = 'июня'
                    elif cellObj.value.month == 7:
                        result_mounth = 'июля'
                    elif cellObj.value.month == 8:
                        result_mounth = 'августа'
                    elif cellObj.value.month == 9:
                        result_mounth = 'сентября'
                    elif cellObj.value.month == 10:
                        result_mounth = 'октября'
                    elif cellObj.value.month == 11:
                        result_mounth = 'ноября'
                    elif cellObj.value.month == 12:
                        result_mounth = 'декабря'
                    date_dict.append(f"{cellObj.value.day} {result_mounth} {cellObj.value.year}")
                except AttributeError:
                    pass
        lessons = {x: y for x in themes_list for y in date_dict}
        # print(date_dict)

        # cell_value = sheet.cell(column,row).value  # cell_value = sheet.cell(строка, столбец)
        # print(cell_value)
        # print(lessons)

    def automatic(pattern, group, lesson, hours):
        for x in range(number_of_lessons):
            doc = DocxTemplate(pattern)
            day = time.localtime()
            d = day[2]
            # n = 64
            context = {'name': name, 'lesson': lesson,
                       'theme': themes_list[x], 'date': date_dict[x], 'number': number_list[x], 'class': group}

            doc.render(context)
            doc.save(f"Техкарты готовые/Занятие_{number_list[x]}.docx")

            # n += 1
            if x % 2 == 0:
                d += 5
            else:
                d += 2

    table(pattern, group)
    automatic(pattern, group, lesson, hours)
    print(f'Добро пожаловать в программу "Генератор технологических карт!"')
    print("Генерируем для Вас файлы")
    for x in range(10):
        print("#####", end='')
        time.sleep(0.5)
    print()
    print("Вот и всё! Готовые файлы были сохранены в папку с программой!")
    os.system(r"explorer.exe Техкарты готовые")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

    sys.exit(app.exec_())